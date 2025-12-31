from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    InputMediaPhoto,
)
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    CallbackQueryHandler,
    MessageHandler,
    ContextTypes,
    filters,
)
from datetime import datetime, timedelta
from openpyxl import load_workbook
import os

# ================== الإعدادات ==================
import os
BOT_TOKEN = os.getenv("BOT_TOKEN")
ADMIN_ID = 5561309338
GROUP_ID = -1003506672561

PAYMENT_NUMBER = "01013008532"
SUB_PRICES = {30: 500, 60: 800, 90: 1200}

EXCEL_FILE = "subscriptions.xlsx"
PROFITS_DIR = "profits"
REVIEWS_DIR = "reviews"

FIXED_START_DATE = datetime(2026, 1, 1)

os.makedirs(PROFITS_DIR, exist_ok=True)
os.makedirs(REVIEWS_DIR, exist_ok=True)

pending_payments = {}

# ================== تصحيح كل التواريخ ==================
def force_fix_dates():
    if not os.path.exists(EXCEL_FILE):
        return

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        days = row[5].value
        if not isinstance(days, int):
            continue

        start = FIXED_START_DATE
        end = start + timedelta(days=days)

        row[3].value = start.strftime("%Y-%m-%d")
        row[4].value = end.strftime("%Y-%m-%d")

    wb.save(EXCEL_FILE)

# ================== القوائم ==================
def main_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🟢 الاشتراك", callback_data="menu_sub")],
        [InlineKeyboardButton("💰 أرباح الأعضاء", callback_data="profits")],
        [InlineKeyboardButton("⭐ تقييمات الأعضاء", callback_data="reviews")],
    ])

def sub_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("30 يوم - 500 جنيه", callback_data="sub_30")],
        [InlineKeyboardButton("60 يوم - 800 جنيه", callback_data="sub_60")],
        [InlineKeyboardButton("90 يوم - 1200 جنيه", callback_data="sub_90")],
        [InlineKeyboardButton("⬅️ رجوع", callback_data="back")],
    ])

# ================== Start ==================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 أهلاً بيك في *مهندس بورصة*\nاختر من القائمة 👇",
        reply_markup=main_menu(),
        parse_mode="Markdown"
    )

# ================== عرض الصور ==================
async def send_images(bot, chat_id, folder, title):
    imgs = [f for f in os.listdir(folder) if f.lower().endswith(("jpg", "png", "jpeg"))]

    if not imgs:
        await bot.send_message(chat_id, "❌ لا توجد صور حالياً")
        return

    await bot.send_message(chat_id, title)

    media = [
        InputMediaPhoto(open(os.path.join(folder, img), "rb"))
        for img in imgs[:10]
    ]

    await bot.send_media_group(chat_id, media)

# ================== الأزرار ==================
async def buttons(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query

    try:
        await q.answer()
    except:
        pass

    uid = q.from_user.id
    data = q.data

    if data == "back":
        await q.message.reply_text("القائمة الرئيسية", reply_markup=main_menu())

    elif data == "menu_sub":
        await q.message.reply_text("اختر مدة الاشتراك 👇", reply_markup=sub_menu())

    elif data.startswith("sub_"):
        days = int(data.split("_")[1])
        pending_payments[uid] = days

        await q.message.reply_text(
            f"🧾 اشتراك {days} يوم\n"
            f"💰 السعر: {SUB_PRICES[days]} جنيه\n\n"
            f"📲 InstaPay / Vodafone Cash\n"
            f"{PAYMENT_NUMBER}\n\n"
            f"📸 ابعت صورة الإيصال"
        )

    elif data == "profits":
        await send_images(context.bot, q.message.chat_id, PROFITS_DIR, "💰 أرباح الأعضاء")

    elif data == "reviews":
        await send_images(context.bot, q.message.chat_id, REVIEWS_DIR, "⭐ تقييمات الأعضاء")

# ================== صورة الإيصال ==================
async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if uid not in pending_payments:
        return

    days = pending_payments[uid]

    kb = InlineKeyboardMarkup([
        [
            InlineKeyboardButton("✅ تأكيد", callback_data=f"ok_{uid}_{days}"),
            InlineKeyboardButton("❌ رفض", callback_data=f"no_{uid}")
        ]
    ])

    user = update.effective_user

    await context.bot.send_photo(
        ADMIN_ID,
        update.message.photo[-1].file_id,
        caption=(
            "📩 طلب اشتراك جديد\n\n"
            f"👤 {user.first_name}\n"
            f"🆔 {uid}\n"
            f"📅 {days} يوم"
        ),
        reply_markup=kb
    )

    await update.message.reply_text("⏳ تم إرسال الإيصال")

# ================== موافقة الأدمن ==================
async def admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    try:
        await q.answer()
    except:
        pass

    if q.from_user.id != ADMIN_ID:
        return

    if q.data.startswith("ok_"):
        _, uid, days = q.data.split("_")
        uid, days = int(uid), int(days)

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        start = FIXED_START_DATE
        end = start + timedelta(days=days)

        ws.append([
            uid,
            "",
            "",
            start.strftime("%Y-%m-%d"),
            end.strftime("%Y-%m-%d"),
            days,
            SUB_PRICES[days],
            SUB_PRICES[days] // (days // 30),
            False,
            False,
            False,
            False
        ])

        wb.save(EXCEL_FILE)

        link = await context.bot.create_chat_invite_link(
            chat_id=GROUP_ID,
            member_limit=1
        )

        await context.bot.send_message(
            uid,
            f"✅ تم تفعيل اشتراكك\n🔗 رابط الدخول:\n{link.invite_link}"
        )

        await q.message.edit_caption("✔️ تم التأكيد")

# ================== تشغيل ==================
print("🔧 Fixing all subscription dates...")
force_fix_dates()

app = ApplicationBuilder().token(BOT_TOKEN).build()

app.add_handler(CommandHandler("start", start))
app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
app.add_handler(CallbackQueryHandler(admin, pattern="^(ok_|no_)"))
app.add_handler(CallbackQueryHandler(buttons))

print("🚀 Bot running safely")
app.run_polling()

